"""
Excel document handler (.xlsx format).

Handles Excel spreadsheets with:
- Text extraction from all sheets
- Cell-by-cell anonymization
- Spreadsheet reconstruction preserving structure
- Support for .xlsx format (OpenXML)
"""

from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime
import io

from .base import BaseDocument, DocumentFormat, DocumentMetadata
from ..logging_config import get_logger
from ..exceptions import DocumentProcessingError

logger = get_logger(__name__)


class ExcelDocument(BaseDocument):
    """
    Excel document handler for .xlsx files.

    Supports:
    - Text extraction from all worksheets
    - Cell-by-cell data extraction
    - Spreadsheet reconstruction with anonymized content
    - Metadata preservation
    - Formula handling (converts to values)

    Note: Requires openpyxl library.
    """

    def __init__(self, file_path: Path):
        """
        Initialize Excel document handler.

        Args:
            file_path: Path to Excel file

        Raises:
            DocumentProcessingError: If Excel document can't be loaded
        """
        self._workbook = None
        self._cell_map = []  # Track cell positions for reconstruction

        super().__init__(file_path)

        # Load document
        self._load_document()

    def _load_document(self):
        """Load Excel document with openpyxl"""
        try:
            from openpyxl import load_workbook

            self._workbook = load_workbook(filename=self.file_path, data_only=True)

            logger.info(
                f"Excel document loaded successfully",
                extra={
                    "extra_fields": {
                        "sheets": len(self._workbook.sheetnames),
                        "file_size": self.get_file_size(),
                    }
                },
            )

        except ImportError:
            raise DocumentProcessingError(
                "openpyxl not installed. Install with: pip install openpyxl",
                {"required_package": "openpyxl"},
            )
        except Exception as e:
            logger.error(f"Failed to load Excel document: {e}", exc_info=True)
            raise DocumentProcessingError(
                f"Failed to load Excel document: {str(e)}",
                {"file_path": str(self.file_path)}
            )

    def _detect_format(self) -> DocumentFormat:
        """Detect format (always EXCEL for this handler)"""
        return DocumentFormat.EXCEL

    def _extract_metadata(self) -> DocumentMetadata:
        """Extract Excel document metadata"""
        file_stats = self.file_path.stat()

        # Extract document properties
        author = None
        title = None
        creation_date = None
        modification_date = None

        try:
            if hasattr(self._workbook, "properties"):
                props = self._workbook.properties
                author = props.creator
                title = props.title
                creation_date = props.created
                modification_date = props.modified

        except Exception as e:
            logger.warning(f"Failed to extract Excel metadata: {e}")

        # Count sheets and cells
        sheet_count = len(self._workbook.sheetnames) if self._workbook else 0

        # Count total cells with data
        total_cells = 0
        if self._workbook:
            for sheet in self._workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            total_cells += 1

        return DocumentMetadata(
            file_name=self.file_path.name,
            file_size=file_stats.st_size,
            format=DocumentFormat.EXCEL,
            author=author,
            title=title,
            creation_date=creation_date,
            modification_date=modification_date,
            is_scanned=False,  # Excel files are digital
            custom={
                "sheets": sheet_count,
                "total_cells": total_cells,
            }
        )

    def extract_text(self) -> str:
        """
        Extract all text from Excel document.

        Extracts from all worksheets, converting each cell to text.
        Formulas are converted to their calculated values.

        Returns:
            Extracted text content with sheet and cell references

        Raises:
            DocumentProcessingError: If extraction fails
        """
        try:
            text_parts = []

            for sheet_name in self._workbook.sheetnames:
                sheet = self._workbook[sheet_name]

                logger.debug(f"Extracting text from sheet: {sheet_name}")

                # Add sheet header
                text_parts.append(f"[Sheet: {sheet_name}]")

                # Extract all cells with data
                for row_idx, row in enumerate(sheet.iter_rows(), start=1):
                    row_texts = []

                    for col_idx, cell in enumerate(row, start=1):
                        if cell.value is not None:
                            cell_text = str(cell.value)
                            row_texts.append(cell_text)

                            # Track cell position for reconstruction
                            self._cell_map.append({
                                "sheet": sheet_name,
                                "row": row_idx,
                                "col": col_idx,
                                "original_value": cell.value,
                                "coordinate": cell.coordinate,
                            })

                    # Add row if it has content
                    if row_texts:
                        text_parts.append(" | ".join(row_texts))

                text_parts.append("")  # Empty line between sheets

            full_text = "\n".join(text_parts)

            logger.info(
                f"Text extraction completed",
                extra={
                    "extra_fields": {
                        "total_sheets": len(self._workbook.sheetnames),
                        "total_cells": len(self._cell_map),
                        "total_length": len(full_text),
                    }
                },
            )

            return full_text

        except Exception as e:
            logger.error(f"Text extraction failed: {e}", exc_info=True)
            raise DocumentProcessingError(
                f"Failed to extract text from Excel document: {str(e)}",
                {"file_path": str(self.file_path)}
            )

    def rebuild(self, anonymized_text: str, detections: List[Dict[str, Any]]) -> bytes:
        """
        Rebuild Excel document with anonymized content.

        Creates a new workbook with anonymized text, attempting to preserve
        sheet structure and cell positions.

        Args:
            anonymized_text: Anonymized text
            detections: List of detections with positions

        Returns:
            Excel document bytes
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            logger.info("Rebuilding Excel document with anonymized content")

            # Create new workbook
            new_wb = Workbook()

            # Remove default sheet
            if "Sheet" in new_wb.sheetnames:
                del new_wb["Sheet"]

            # Parse anonymized text back into structure
            lines = anonymized_text.split("\n")
            current_sheet = None
            current_sheet_obj = None

            for line in lines:
                if not line.strip():
                    continue

                # Check if it's a sheet header
                if line.startswith("[Sheet:"):
                    sheet_name = line.replace("[Sheet:", "").replace("]", "").strip()
                    current_sheet = sheet_name
                    current_sheet_obj = new_wb.create_sheet(title=sheet_name)
                    logger.debug(f"Creating sheet: {sheet_name}")
                    continue

                # Parse row data
                if current_sheet_obj is not None:
                    cells = line.split(" | ")

                    # Find next available row
                    row_idx = current_sheet_obj.max_row + 1
                    if current_sheet_obj.max_row == 1 and current_sheet_obj['A1'].value is None:
                        row_idx = 1

                    # Write cells
                    for col_idx, cell_value in enumerate(cells, start=1):
                        cell = current_sheet_obj.cell(row=row_idx, column=col_idx)
                        cell.value = cell_value

            # Apply basic styling (optional)
            for sheet in new_wb.worksheets:
                # Make first row bold (header style)
                for cell in sheet[1]:
                    if cell.value:
                        cell.font = Font(bold=True)

            # Save to bytes
            buffer = io.BytesIO()
            new_wb.save(buffer)
            excel_bytes = buffer.getvalue()
            buffer.close()

            logger.info(
                "Excel document rebuilt successfully",
                extra={"extra_fields": {"output_size": len(excel_bytes)}},
            )

            return excel_bytes

        except Exception as e:
            logger.error(f"Excel document rebuild failed: {e}", exc_info=True)
            raise DocumentProcessingError(
                f"Failed to rebuild Excel document: {str(e)}",
                {"file_path": str(self.file_path)}
            )

    def get_sheet_names(self) -> List[str]:
        """
        Get list of sheet names.

        Returns:
            List of sheet names
        """
        if hasattr(self, "_workbook") and self._workbook:
            return self._workbook.sheetnames
        return []

    def get_sheet_count(self) -> int:
        """
        Get number of sheets.

        Returns:
            Sheet count
        """
        if hasattr(self, "_workbook") and self._workbook:
            return len(self._workbook.sheetnames)
        return 0

    def get_cell_count(self) -> int:
        """
        Get total number of cells with data.

        Returns:
            Cell count
        """
        return len(self._cell_map)

    def close(self):
        """Close workbook resources"""
        if hasattr(self, "_workbook") and self._workbook:
            self._workbook.close()
            logger.debug("Excel workbook closed")

    def __del__(self):
        """Cleanup on deletion"""
        self.close()
